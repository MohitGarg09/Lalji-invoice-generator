from io import BytesIO
from typing import Iterable
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment
from django.db import transaction
from django.conf import settings
import os
import tempfile
import shutil
import time
import logging
from .models import Invoice, InvoicePDFRecord

logger = logging.getLogger(__name__)

def _safe_save_workbook(wb, path: str):
    """
    Save workbook safely to avoid duplicate copies like (1), (2) when file is open.
    Writes to a temporary file first then atomically moves it into place.
    Retries a few times if the target file is locked (PermissionError / OSError).
    Returns True on success, False on failure.
    """
    temp_dir = os.path.dirname(os.path.abspath(path)) or os.getcwd()
    try:
        fd, tmp_path = tempfile.mkstemp(suffix=".xlsx", dir=temp_dir)
        os.close(fd)
    except Exception:
        # Fallback to simple .tmp next to target path
        tmp_path = f"{path}.tmp"

    max_retries = 3
    for attempt in range(1, max_retries + 1):
        try:
            # Save workbook to temp path
            wb.save(tmp_path)
            # Move (atomic on many platforms) temp file to final path
            shutil.move(tmp_path, path)
            return True
        except (PermissionError, OSError) as e:
            logger.warning("Attempt %d/%d: file %s locked or inaccessible (%s). Retrying...", attempt, max_retries, path, e)
            time.sleep(1)
            continue
        except Exception as e:
            logger.exception("Failed to save workbook to %s: %s", path, e)
            # Ensure temp file cleaned up if present
            try:
                if os.path.exists(tmp_path):
                    os.remove(tmp_path)
            except Exception:
                pass
            return False

    logger.error("Could not save workbook to %s after %d attempts (file may be locked).", path, max_retries)
    # Clean up temp if still present
    try:
        if os.path.exists(tmp_path):
            os.remove(tmp_path)
    except Exception:
        pass
    return False


INVOICE_RECORDS_HEADERS = [
    "invoice_id", "customer_name", "created_at", "bill_type",
    "subtotal", "discount_percent", "gst_amount", "total_amount",
    "pdf_link", "notes"
]

# Desktop fallback path (used only if settings.INVOICE_MASTER_EXCEL isn't set)
DESKTOP_PATH = os.path.join(os.path.expanduser("~"), "Desktop")
if not os.path.exists(DESKTOP_PATH):
    try:
        os.makedirs(DESKTOP_PATH, exist_ok=True)
    except Exception:
        # fallback to settings.BASE_DIR (should exist)
        DESKTOP_PATH = str(getattr(settings, "BASE_DIR", os.getcwd()))

# MASTER_EXCEL_PATH uses the configured setting if present, else falls back to Desktop
MASTER_EXCEL_PATH = getattr(settings, "INVOICE_MASTER_EXCEL",
                            os.path.join(DESKTOP_PATH, 'invoice_excel.xlsx'))


def _ensure_parent_dir(path: str):
    parent_dir = os.path.dirname(path)
    if parent_dir and not os.path.exists(parent_dir):
        try:
            os.makedirs(parent_dir, exist_ok=True)
        except Exception as e:
            logger.warning("Could not create parent directory %s: %s", parent_dir, e)


def _get_or_create_master_excel():
    """Get existing master Excel or create a new one with headers."""
    if os.path.exists(MASTER_EXCEL_PATH):
        try:
            wb = load_workbook(MASTER_EXCEL_PATH)
            ws = wb.active
            return wb, ws
        except Exception as e:
            logger.warning("Could not load existing master Excel %s (%s). Will create a new workbook.", MASTER_EXCEL_PATH, e)

    # create new workbook and write headers
    wb = Workbook()
    ws = wb.active
    ws.title = "Invoice Records"

    # Add headers with formatting
    ws.append(INVOICE_RECORDS_HEADERS)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Ensure parent dir exists and save safely
    _ensure_parent_dir(MASTER_EXCEL_PATH)
    saved = _safe_save_workbook(wb, MASTER_EXCEL_PATH)
    if not saved:
        logger.error("Failed to create initial master Excel at %s", MASTER_EXCEL_PATH)

    return wb, ws


def add_invoice_record_to_master(invoice: Invoice, pdf_link: str = "", notes: str = "") -> None:
    """Append or update a single invoice record in the master Excel file."""
    wb, ws = _get_or_create_master_excel()

    # Check if invoice already exists and update instead of duplicate
    invoice_exists = False
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=False), start=2):
        if row and row[0].value == invoice.id:
            # Update existing row - set values by column index
            try:
                ws.cell(row=row_idx, column=1, value=invoice.id)
                ws.cell(row=row_idx, column=2, value=invoice.customer_name or "-")
                ws.cell(row=row_idx, column=3, value=invoice.created_at.strftime("%Y-%m-%d %H:%M"))
                ws.cell(row=row_idx, column=4, value=invoice.bill_type or "-")
                ws.cell(row=row_idx, column=5, value=float(invoice.total))
                ws.cell(row=row_idx, column=6, value=float(invoice.discount_percent))
                ws.cell(row=row_idx, column=7, value=float(invoice.gst_amount))
                ws.cell(row=row_idx, column=8, value=float(invoice.total_with_gst))
                ws.cell(row=row_idx, column=9, value=pdf_link)
                ws.cell(row=row_idx, column=10, value=notes)
            except Exception as e:
                logger.exception("Error updating row for invoice %s: %s", invoice.id, e)
            invoice_exists = True
            break

    # If invoice doesn't exist, append as new row
    if not invoice_exists:
        try:
            ws.append([
                invoice.id,
                invoice.customer_name or "-",
                invoice.created_at.strftime("%Y-%m-%d %H:%M"),
                invoice.bill_type or "-",
                float(invoice.total),
                float(invoice.discount_percent),
                float(invoice.gst_amount),
                float(invoice.total_with_gst),
                pdf_link,
                notes,
            ])
        except Exception as e:
            logger.exception("Error appending row for invoice %s: %s", invoice.id, e)

    # Auto-adjust column widths
    try:
        for col in ws.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except Exception:
                    pass
            ws.column_dimensions[col_letter].width = max_length + 2
    except Exception as e:
        logger.exception("Error auto-adjusting column widths: %s", e)

    # Ensure parent dir exists and save workbook safely
    _ensure_parent_dir(MASTER_EXCEL_PATH)
    saved = _safe_save_workbook(wb, MASTER_EXCEL_PATH)
    if not saved:
        logger.error("Failed to save master Excel after adding/updating invoice %s", invoice.id)


def export_invoice_records_to_excel(invoices: Iterable[Invoice]) -> bytes:
    """Return the master Excel file bytes (single file)."""
    if not os.path.exists(MASTER_EXCEL_PATH):
        _get_or_create_master_excel()

    try:
        with open(MASTER_EXCEL_PATH, 'rb') as f:
            return f.read()
    except Exception as e:
        logger.exception("Failed to read master Excel file %s: %s", MASTER_EXCEL_PATH, e)
        # Return an empty workbook bytes as fallback
        wb = Workbook()
        bio = BytesIO()
        wb.save(bio)
        return bio.getvalue()


@transaction.atomic
def import_invoice_records_from_excel(excel_bytes: bytes) -> int:
    """Import invoice records with PDF links from Excel file into DB."""
    wb = load_workbook(filename=BytesIO(excel_bytes))
    ws = wb.active

    # Read header row and compare case-insensitively
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    expected = [h.lower() for h in INVOICE_RECORDS_HEADERS]
    actual = [(h.lower() if isinstance(h, str) else h) for h in headers[:len(INVOICE_RECORDS_HEADERS)]]

    if actual != expected:
        raise ValueError("Invalid header row for invoice records import. Expected: " + ",".join(INVOICE_RECORDS_HEADERS))

    imported_count = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(v is None for v in row):
            continue

        invoice_id = row[0]
        pdf_link = row[8] if len(row) > 8 else ""
        notes = row[9] if len(row) > 9 else ""

        try:
            invoice = Invoice.objects.get(id=invoice_id)
            pdf_record, _created = InvoicePDFRecord.objects.update_or_create(
                invoice=invoice,
                defaults={
                    'pdf_file_path': str(pdf_link).strip() if pdf_link else "",
                    'notes': str(notes).strip() if notes else "",
                }
            )
            imported_count += 1
        except Invoice.DoesNotExist:
            raise ValueError(f"Invoice with ID {invoice_id} not found")

    return imported_count


def add_pdf_link_to_invoice(invoice: Invoice, pdf_file_path: str, notes: str = "") -> InvoicePDFRecord:
    """Add or update PDF link record for an invoice and append to master Excel."""
    pdf_record, _created = InvoicePDFRecord.objects.update_or_create(
        invoice=invoice,
        defaults={
            'pdf_file_path': pdf_file_path,
            'notes': notes,
        }
    )

    # Append / update entry in the master Excel file
    try:
        add_invoice_record_to_master(invoice, pdf_file_path, notes)
    except Exception as e:
        logger.exception("Failed to add invoice record to master excel for invoice %s: %s", invoice.id, e)

    return pdf_record


def get_invoice_pdf_link(invoice: Invoice) -> str:
    """Get PDF link for an invoice."""
    pdf_record = InvoicePDFRecord.objects.filter(invoice=invoice).first()
    return pdf_record.pdf_file_path if pdf_record else None


def export_single_invoice_record_to_excel(invoice: Invoice) -> bytes:
    """
    Add or update the given invoice row in the master Excel (MASTER_EXCEL_PATH)
    and return the master Excel bytes. This avoids creating separate invoice_{id}.xlsx files.
    """
    wb, ws = _get_or_create_master_excel()

    pdf_rec = InvoicePDFRecord.objects.filter(invoice=invoice).first()

    row_values = [
        invoice.id,
        invoice.customer_name or "-",
        invoice.created_at.strftime("%Y-%m-%d %H:%M"),
        invoice.bill_type or "-",
        float(invoice.total),
        float(invoice.discount_percent),
        float(invoice.gst_amount),
        float(invoice.total_with_gst),
        pdf_rec.pdf_file_path if pdf_rec else "",
        pdf_rec.notes if pdf_rec else "",
    ]

    updated = False
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=False), start=2):
        if row and row[0].value == invoice.id:
            for col_idx, value in enumerate(row_values, start=1):
                ws.cell(row=row_idx, column=col_idx, value=value)
            updated = True
            break

    if not updated:
        ws.append(row_values)

    # Auto-adjust column widths
    try:
        for col in ws.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except Exception:
                    pass
            ws.column_dimensions[col_letter].width = max_length + 2
    except Exception as e:
        logger.exception("Error auto-adjusting column widths: %s", e)

    _ensure_parent_dir(MASTER_EXCEL_PATH)
    saved = _safe_save_workbook(wb, MASTER_EXCEL_PATH)
    if not saved:
        logger.error("Failed to save master Excel when exporting single invoice %s", invoice.id)

    try:
        with open(MASTER_EXCEL_PATH, 'rb') as f:
            return f.read()
    except Exception as e:
        logger.exception("Failed to read master Excel file after saving: %s", e)
        bio = BytesIO()
        wb.save(bio)
        return bio.getvalue()


def get_master_excel_as_bytes() -> bytes:
    """Return the master Excel file as bytes for download."""
    if not os.path.exists(MASTER_EXCEL_PATH):
        _get_or_create_master_excel()

    try:
        with open(MASTER_EXCEL_PATH, 'rb') as f:
            return f.read()
    except Exception as e:
        logger.exception("Failed to read master Excel file %s: %s", MASTER_EXCEL_PATH, e)
        wb = Workbook()
        bio = BytesIO()
        wb.save(bio)
        return bio.getvalue()
